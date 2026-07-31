#!/usr/bin/env python3
"""
Runs bills flagged by Qwen (07_31 run) through GPT-4o, Claude Sonnet, AND Claude Opus.
All three models use the same prompt: 8_llm_categorize_VT27_bills_prompt_07_31.md
(the original full-detail 1-step prompt, not the simplified step-2 prompt).

Input:
  - ../llm_bill_VT27_merged_2010sample5000_07_31_qwen.csv   (Qwen results, 140 flagged)
  - D:/bill_summary/SAMPLE_5000_YEAR_2010_2019/llm_input_metadata_no_prefilter.csv
  - D:/bill_summary/SAMPLE_5000_YEAR_2010_2019/llm_input_texts_no_prefilter/
  - ../HPC/8_llm_categorize_VT27_bills_prompt_07_31.md

Output:
  - OUTPUT/llm_bill_VT27_compare_07_31_3model.csv
  - OUTPUT/compare_07_31_3model_checkpoint.csv  (auto-deleted on completion)
  - ../manual_review_07_31_3model.xlsx           (any-model disagreements)
  - OUTPUT/api_cost_07_31_3model.txt
"""

import csv
import json
import os
import time
from datetime import datetime
from pathlib import Path

import anthropic
from openai import OpenAI
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

csv.field_size_limit(10_000_000)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
KEY_FILE   = Path(r"D:\Dropbox\fengheliu\tools\keys.txt")
SCRIPT_DIR = Path(__file__).resolve().parent
REPO_DIR   = SCRIPT_DIR.parent

QWEN_CSV    = REPO_DIR / "llm_bill_VT27_merged_2010sample5000_07_31_qwen.csv"
META_CSV    = Path(r"D:\bill_summary\SAMPLE_5000_YEAR_2010_2019") / "llm_input_metadata_no_prefilter.csv"
TEXT_DIR    = Path(r"D:\bill_summary\SAMPLE_5000_YEAR_2010_2019") / "llm_input_texts_no_prefilter"
PROMPT_MD   = REPO_DIR / "HPC" / "8_llm_categorize_VT27_bills_prompt_07_31.md"
OUTPUT_CSV  = SCRIPT_DIR / "OUTPUT" / "llm_bill_VT27_compare_07_31_3model.csv"
CHECKPOINT  = SCRIPT_DIR / "OUTPUT" / "compare_07_31_3model_checkpoint.csv"
REVIEW_XLS  = REPO_DIR / "manual_review_07_31_3model.xlsx"
COST_FILE   = SCRIPT_DIR / "OUTPUT" / "api_cost_07_31_3model.txt"

GPT_MODEL    = "gpt-4o"
SONNET_MODEL = "claude-sonnet-4-5"
OPUS_MODEL   = "claude-opus-4-6"
CHECKPOINT_N = 20

# ---------------------------------------------------------------------------
# Load API keys
# ---------------------------------------------------------------------------
def load_openai_key(key_file: Path) -> str:
    if key_file.exists():
        for line in key_file.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line.startswith("sk-") and "ant" not in line:
                return line
            if ": sk-" in line and "ant" not in line:
                return line.split(": ", 1)[1].strip()
    key = os.environ.get("OPENAI_API_KEY", "")
    if key:
        return key
    raise FileNotFoundError("No OpenAI key found.")

def load_anthropic_key(key_file: Path) -> str:
    if key_file.exists():
        for line in key_file.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line.startswith("sk-ant-"):
                return line
            if ": sk-ant-" in line:
                return line.split(": ", 1)[1].strip()
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if key:
        return key
    raise FileNotFoundError("No Anthropic key found.")

gpt_client    = OpenAI(api_key=load_openai_key(KEY_FILE))
claude_client = anthropic.Anthropic(api_key=load_anthropic_key(KEY_FILE))

# ---------------------------------------------------------------------------
# Build system prompt
# ---------------------------------------------------------------------------
def build_system_prompt(md_path: Path) -> str:
    text = md_path.read_text(encoding="utf-8")
    before = text.split("## Bill Information")[0].rstrip()
    output = "## Output" + text.split("## Output")[1].rstrip()
    return before + "\n\n" + output

SYSTEM_PROMPT = build_system_prompt(PROMPT_MD)

# ---------------------------------------------------------------------------
# Load Qwen-flagged bills
# ---------------------------------------------------------------------------
qwen_rows = {}
with open(QWEN_CSV, encoding="utf-8") as f:
    for r in csv.DictReader(f):
        if r["category_ids"].strip() in ("1", "1.0"):
            qwen_rows[r["bill_id"]] = r

print(f"Step-1 flagged bills: {len(qwen_rows)}")

# ---------------------------------------------------------------------------
# Load metadata
# ---------------------------------------------------------------------------
meta = {}
with open(META_CSV, encoding="utf-8") as f:
    for r in csv.DictReader(f):
        if r["bill_id"] in qwen_rows:
            meta[r["bill_id"]] = r

print(f"Metadata found: {len(meta)} / {len(qwen_rows)}")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def build_user_message(bill_id: str, row: dict) -> str:
    text_file = TEXT_DIR / f"{bill_id}_summary.txt"
    summary   = text_file.read_text(encoding="utf-8").strip() if text_file.exists() else row.get("summary", "")
    return (
        f"## Bill Information\n\n"
        f"**Official title:** {row['official_title']}\n"
        f"**Short title:** {row['short_title']}\n"
        f"**Policy area:** {row['policy_area']}\n"
        f"**Subjects:** {row['subjects']}\n"
        f"**Summary:**\n{summary}"
    )

def strip_code_fence(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1] if "\n" in text else text
        if text.endswith("```"):
            text = text[:-3].rstrip()
    return text

def parse_response(content: str) -> tuple[str, str, str, str]:
    parsed        = json.loads(content)
    cats          = parsed.get("categories", [])
    bill_type     = parsed.get("bill_type", "")
    bill_category = parsed.get("bill_category", "")
    reason        = parsed.get("reason", "")
    cats_str      = ";".join(str(c) for c in cats) if cats else ""
    return cats_str, bill_type, bill_category, reason

out_fields = [
    "bill_id", "official_title", "short_title", "policy_area", "subjects",
    "category_qwen",   "bill_type_qwen",   "bill_category_qwen",   "reason_qwen",
    "category_4o",     "bill_type_4o",     "bill_category_4o",     "reason_4o",
    "category_sonnet", "bill_type_sonnet", "bill_category_sonnet", "reason_sonnet",
    "category_opus",   "bill_type_opus",   "bill_category_opus",   "reason_opus",
]

# ---------------------------------------------------------------------------
# Cost tracking
# ---------------------------------------------------------------------------
gpt_input_tokens = gpt_output_tokens = 0
sonnet_input_tokens = sonnet_output_tokens = 0
opus_input_tokens = opus_output_tokens = 0

# ---------------------------------------------------------------------------
# Load checkpoint
# ---------------------------------------------------------------------------
done: dict[str, dict] = {}
if CHECKPOINT.exists():
    with open(CHECKPOINT, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            done[r["bill_id"]] = r
    print(f"Resumed from checkpoint: {len(done)} bills already done.")

results = list(done.values())
new_since_checkpoint = 0

# ---------------------------------------------------------------------------
# Main loop
# ---------------------------------------------------------------------------
bill_list = list(qwen_rows.items())

for idx, (bill_id, q_row) in enumerate(bill_list, 1):
    if bill_id in done:
        continue

    m = meta.get(bill_id, {})
    if not m:
        print(f"[{idx}/{len(bill_list)}] {bill_id} — no metadata, skipping")
        continue

    print(f"\n[{idx}/{len(bill_list)}] {bill_id}")
    user_msg = build_user_message(bill_id, m)

    # --- GPT-4o ---
    gpt_cats = gpt_bt = gpt_bc = gpt_reason = "API error"
    for attempt in range(1, 4):
        try:
            resp = gpt_client.chat.completions.create(
                model=GPT_MODEL,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user",   "content": user_msg},
                ],
                temperature=0.0,
                max_tokens=400,
                response_format={"type": "json_object"},
            )
            gpt_cats, gpt_bt, gpt_bc, gpt_reason = parse_response(resp.choices[0].message.content)
            gpt_input_tokens  += resp.usage.prompt_tokens
            gpt_output_tokens += resp.usage.completion_tokens
            print(f"  GPT-4o:  [{gpt_cats or '-'}] [{gpt_bt or '-'}] [{gpt_bc or '-'}] -- {gpt_reason[:80]}")
            break
        except Exception as e:
            print(f"  GPT-4o retry {attempt}: {e}")
            time.sleep(2 ** attempt)

    # --- Claude Sonnet ---
    sn_cats = sn_bt = sn_bc = sn_reason = "API error"
    for attempt in range(1, 4):
        try:
            resp = claude_client.messages.create(
                model=SONNET_MODEL,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": user_msg}],
                temperature=0.0,
                max_tokens=400,
            )
            content = strip_code_fence(resp.content[0].text)
            sn_cats, sn_bt, sn_bc, sn_reason = parse_response(content)
            sonnet_input_tokens  += resp.usage.input_tokens
            sonnet_output_tokens += resp.usage.output_tokens
            print(f"  Sonnet:  [{sn_cats or '-'}] [{sn_bt or '-'}] [{sn_bc or '-'}] -- {sn_reason[:80]}")
            break
        except Exception as e:
            print(f"  Sonnet retry {attempt}: {e}")
            time.sleep(2 ** attempt)

    # --- Claude Opus ---
    op_cats = op_bt = op_bc = op_reason = "API error"
    for attempt in range(1, 4):
        try:
            resp = claude_client.messages.create(
                model=OPUS_MODEL,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": user_msg}],
                temperature=0.0,
                max_tokens=400,
            )
            content = strip_code_fence(resp.content[0].text)
            op_cats, op_bt, op_bc, op_reason = parse_response(content)
            opus_input_tokens  += resp.usage.input_tokens
            opus_output_tokens += resp.usage.output_tokens
            print(f"  Opus:    [{op_cats or '-'}] [{op_bt or '-'}] [{op_bc or '-'}] -- {op_reason[:80]}")
            break
        except Exception as e:
            print(f"  Opus retry {attempt}: {e}")
            time.sleep(2 ** attempt)

    row_out = {
        "bill_id":               bill_id,
        "official_title":        m.get("official_title", ""),
        "short_title":           m.get("short_title", ""),
        "policy_area":           m.get("policy_area", ""),
        "subjects":              m.get("subjects", ""),
        "category_qwen":         q_row.get("category_ids", ""),
        "bill_type_qwen":        q_row.get("bill_type", ""),
        "bill_category_qwen":    q_row.get("bill_category", ""),
        "reason_qwen":           q_row.get("reason", ""),
        "category_4o":           gpt_cats,
        "bill_type_4o":          gpt_bt,
        "bill_category_4o":      gpt_bc,
        "reason_4o":             gpt_reason,
        "category_sonnet":       sn_cats,
        "bill_type_sonnet":      sn_bt,
        "bill_category_sonnet":  sn_bc,
        "reason_sonnet":         sn_reason,
        "category_opus":         op_cats,
        "bill_type_opus":        op_bt,
        "bill_category_opus":    op_bc,
        "reason_opus":           op_reason,
    }
    results.append(row_out)
    new_since_checkpoint += 1

    if new_since_checkpoint >= CHECKPOINT_N:
        with open(CHECKPOINT, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=out_fields)
            writer.writeheader()
            writer.writerows(results)
        print(f"  [Checkpoint saved: {len(results)} rows]")
        new_since_checkpoint = 0

# ---------------------------------------------------------------------------
# Save final CSV
# ---------------------------------------------------------------------------
OUTPUT_CSV.parent.mkdir(exist_ok=True)
with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=out_fields)
    writer.writeheader()
    writer.writerows(results)

print(f"\nDone. {len(results)} rows saved to {OUTPUT_CSV}")

if CHECKPOINT.exists():
    CHECKPOINT.unlink()

# ---------------------------------------------------------------------------
# Summary table
# ---------------------------------------------------------------------------
def is_flagged(val):
    return str(val).strip() not in ("", "0", "0.0", "[]", "API error")

print("\n=== SUMMARY ===")
print(f"{'Bill':<18} {'Qwen':>5} {'4o':>4} {'Sonnet':>7} {'Opus':>6}")
print("-" * 46)
for r in results:
    q  = "Y" if is_flagged(r["category_qwen"])   else "N"
    o  = "Y" if is_flagged(r["category_4o"])      else "N"
    s  = "Y" if is_flagged(r["category_sonnet"])  else "N"
    op = "Y" if is_flagged(r["category_opus"])    else "N"
    flag = "  *" if not (o == s == op) else ""
    print(f"{r['bill_id']:<18} {q:>5} {o:>4} {s:>7} {op:>6}{flag}")

gpt_total    = sum(1 for r in results if is_flagged(r["category_4o"]))
sonnet_total = sum(1 for r in results if is_flagged(r["category_sonnet"]))
opus_total   = sum(1 for r in results if is_flagged(r["category_opus"]))
all_agree    = sum(
    1 for r in results
    if is_flagged(r["category_4o"]) == is_flagged(r["category_sonnet"]) == is_flagged(r["category_opus"])
)

gpt_opus_disagree = sum(1 for r in results if is_flagged(r["category_4o"]) != is_flagged(r["category_opus"]))
print(f"\nAll 3 agree:       {all_agree}/{len(results)}")
print(f"GPT vs Opus disagree: {gpt_opus_disagree}/{len(results)}")
print(f"Qwen flagged:   {len(results)}")
print(f"4o flagged:     {gpt_total}")
print(f"Sonnet flagged: {sonnet_total}")
print(f"Opus flagged:   {opus_total}")

# ---------------------------------------------------------------------------
# Generate manual review xlsx (any-model disagreements)
# ---------------------------------------------------------------------------
disagree_rows = [
    r for r in results
    if is_flagged(r["category_4o"]) != is_flagged(r["category_opus"])
]

print(f"\nGPT-4o vs Opus disagreement: {len(disagree_rows)} bills -> {REVIEW_XLS}")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Manual Review"

header_font        = Font(bold=True, color="FFFFFF")
header_fill_4o     = PatternFill("solid", fgColor="1F4E79")   # dark blue   = GPT-4o
header_fill_sonnet = PatternFill("solid", fgColor="375623")   # dark green  = Sonnet
header_fill_opus   = PatternFill("solid", fgColor="3B1A6B")   # dark purple = Opus
header_fill_meta   = PatternFill("solid", fgColor="404040")   # dark grey   = metadata
header_fill_manual = PatternFill("solid", fgColor="7B2C2C")   # dark red    = manual columns

fill_4o_yes     = PatternFill("solid", fgColor="BDD7EE")   # light blue
fill_sonnet_yes = PatternFill("solid", fgColor="C6E0B4")   # light green
fill_opus_yes   = PatternFill("solid", fgColor="D9C9F0")   # light purple

columns = [
    # (header, field_key, width, fill_group)
    ("bill_id",          "bill_id",               14, "meta"),
    ("official_title",   "official_title",         60, "meta"),
    ("short_title",      "short_title",            30, "meta"),
    ("policy_area",      "policy_area",            20, "meta"),
    ("4o_cat",           "category_4o",             7, "4o"),
    ("4o_type",          "bill_type_4o",           14, "4o"),
    ("4o_bill_cat",      "bill_category_4o",       12, "4o"),
    ("4o_reason",        "reason_4o",              60, "4o"),
    ("sonnet_cat",       "category_sonnet",         9, "sonnet"),
    ("sonnet_type",      "bill_type_sonnet",       14, "sonnet"),
    ("sonnet_bill_cat",  "bill_category_sonnet",   12, "sonnet"),
    ("sonnet_reason",    "reason_sonnet",           60, "sonnet"),
    ("opus_cat",         "category_opus",           8, "opus"),
    ("opus_type",        "bill_type_opus",         14, "opus"),
    ("opus_bill_cat",    "bill_category_opus",     12, "opus"),
    ("opus_reason",      "reason_opus",            60, "opus"),
    ("qwen_cat",         "category_qwen",           8, "meta"),
    ("qwen_reason",      "reason_qwen",            50, "meta"),
    ("manual_decision",  None,                     16, "manual"),
    ("notes",            None,                     40, "manual"),
]

fill_map = {
    "meta":   header_fill_meta,
    "4o":     header_fill_4o,
    "sonnet": header_fill_sonnet,
    "opus":   header_fill_opus,
    "manual": header_fill_manual,
}

# Write header
for col_idx, (header, _, width, group) in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = fill_map[group]
    cell.alignment = Alignment(wrap_text=False, vertical="center")
    ws.column_dimensions[cell.column_letter].width = width

ws.row_dimensions[1].height = 20

# Write data rows
for row_idx, r in enumerate(disagree_rows, 2):
    flagged_4o     = is_flagged(r["category_4o"])
    flagged_sonnet = is_flagged(r["category_sonnet"])
    flagged_opus   = is_flagged(r["category_opus"])

    for col_idx, (_, field, _, group) in enumerate(columns, 1):
        value = r.get(field, "") if field else ""
        cell  = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if group == "4o" and flagged_4o:
            cell.fill = fill_4o_yes
        elif group == "sonnet" and flagged_sonnet:
            cell.fill = fill_sonnet_yes
        elif group == "opus" and flagged_opus:
            cell.fill = fill_opus_yes

ws.freeze_panes = "A2"
wb.save(REVIEW_XLS)
print(f"Manual review workbook saved: {REVIEW_XLS}")

# ---------------------------------------------------------------------------
# Save cost file
# ---------------------------------------------------------------------------
GPT_INPUT_PRICE    = 2.50 / 1_000_000
GPT_OUTPUT_PRICE   = 10.00 / 1_000_000
SONNET_INPUT_PRICE = 3.00 / 1_000_000
SONNET_OUTPUT_PRICE= 15.00 / 1_000_000
OPUS_INPUT_PRICE   = 15.00 / 1_000_000
OPUS_OUTPUT_PRICE  = 75.00 / 1_000_000

gpt_cost    = gpt_input_tokens * GPT_INPUT_PRICE    + gpt_output_tokens * GPT_OUTPUT_PRICE
sonnet_cost = sonnet_input_tokens * SONNET_INPUT_PRICE + sonnet_output_tokens * SONNET_OUTPUT_PRICE
opus_cost   = opus_input_tokens * OPUS_INPUT_PRICE   + opus_output_tokens * OPUS_OUTPUT_PRICE
total_cost  = gpt_cost + sonnet_cost + opus_cost

cost_lines = f"""API Cost - compare_bills_07_31_3model.py
==========================================
Run completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} (local time)
Script:        GPT_identify_bills/compare_bills_07_31_3model.py
Prompt:        HPC/8_llm_categorize_VT27_bills_prompt_07_31.md
Input pool:    llm_bill_VT27_merged_2010sample5000_07_31_qwen.csv (140 Qwen-flagged)
Bills processed: {len(results)}

--- TRUE Token Counts (from API response.usage) ---

GPT-4o ({GPT_MODEL}):
  Input tokens:   {gpt_input_tokens:,}
  Output tokens:  {gpt_output_tokens:,}

Claude Sonnet ({SONNET_MODEL}):
  Input tokens:   {sonnet_input_tokens:,}
  Output tokens:  {sonnet_output_tokens:,}

Claude Opus ({OPUS_MODEL}):
  Input tokens:   {opus_input_tokens:,}
  Output tokens:  {opus_output_tokens:,}

--- Pricing Used ---

GPT-4o:           $2.50 / 1M input,   $10.00 / 1M output
Claude Sonnet:    $3.00 / 1M input,   $15.00 / 1M output
Claude Opus:     $15.00 / 1M input,   $75.00 / 1M output

--- Cost Breakdown ---

GPT-4o:
  Input cost:   ${gpt_input_tokens * GPT_INPUT_PRICE:.4f}
  Output cost:  ${gpt_output_tokens * GPT_OUTPUT_PRICE:.4f}
  Subtotal:     ${gpt_cost:.4f}

Claude Sonnet ({SONNET_MODEL}):
  Input cost:   ${sonnet_input_tokens * SONNET_INPUT_PRICE:.4f}
  Output cost:  ${sonnet_output_tokens * SONNET_OUTPUT_PRICE:.4f}
  Subtotal:     ${sonnet_cost:.4f}

Claude Opus ({OPUS_MODEL}):
  Input cost:   ${opus_input_tokens * OPUS_INPUT_PRICE:.4f}
  Output cost:  ${opus_output_tokens * OPUS_OUTPUT_PRICE:.4f}
  Subtotal:     ${opus_cost:.4f}

TOTAL:          ${total_cost:.4f}

"""

COST_FILE.write_text(cost_lines, encoding="utf-8")
print(f"\nCost file saved: {COST_FILE}")
print(f"Estimated total cost: ${total_cost:.4f}")
