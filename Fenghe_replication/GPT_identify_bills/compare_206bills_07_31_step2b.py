#!/usr/bin/env python3
"""
Step-2 pass (variant b): same as compare_206bills_07_31_step2.py but the
Qwen step-1 reason is NOT passed to the LLMs as input.  The models only see
the bill metadata and summary, plus the detailed classification rules.
The step-1 reason is still recorded in the output CSV for reference.

Input:
  - ../llm_bill_VT27_merged_2010sample5000_07_31_step1.csv  (206 flagged)
  - D:/bill_summary/SAMPLE_5000_YEAR_2010_2019/llm_input_metadata_no_prefilter.csv
  - D:/bill_summary/SAMPLE_5000_YEAR_2010_2019/llm_input_texts_no_prefilter/
  - ../HPC/prompt_step2_gpt.md

Output:
  - OUTPUT/llm_bill_VT27_compare_07_31_step2b.csv
  - OUTPUT/compare_07_31_step2b_checkpoint.csv  (auto-deleted on completion)
  - ../manual_review_07_31_step2b.xlsx          (any model disagreement)
  - OUTPUT/api_cost_07_31_step2b.txt            (TRUE cost from API usage)
"""

import csv
import json
import os
import time
from pathlib import Path

import anthropic
from openai import OpenAI
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

csv.field_size_limit(10_000_000)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
KEY_FILE   = Path(r"D:\Dropbox\fengheliu\tools\keys.txt")
SCRIPT_DIR = Path(__file__).resolve().parent
REPO_DIR   = SCRIPT_DIR.parent

STEP1_CSV  = REPO_DIR / "llm_bill_VT27_merged_2010sample5000_07_31_step1.csv"
META_CSV   = Path(r"D:\bill_summary\SAMPLE_5000_YEAR_2010_2019") / "llm_input_metadata_no_prefilter.csv"
TEXT_DIR   = Path(r"D:\bill_summary\SAMPLE_5000_YEAR_2010_2019") / "llm_input_texts_no_prefilter"
PROMPT_MD  = REPO_DIR / "HPC" / "prompt_step2_gpt.md"
OUTPUT_CSV = SCRIPT_DIR / "OUTPUT" / "llm_bill_VT27_compare_07_31_step2b.csv"
CHECKPOINT = SCRIPT_DIR / "OUTPUT" / "compare_07_31_step2b_checkpoint.csv"
REVIEW_XLS = REPO_DIR / "manual_review_07_31_step2b.xlsx"
COST_TXT   = SCRIPT_DIR / "OUTPUT" / "api_cost_07_31_step2b.txt"

GPT_MODEL    = "gpt-4o"
SONNET_MODEL = "claude-sonnet-4-5"
OPUS_MODEL   = "claude-opus-4-6"
CHECKPOINT_N = 20

# ---------------------------------------------------------------------------
# Load API keys
# ---------------------------------------------------------------------------
def load_openai_key(key_file: Path) -> str:
    if key_file.exists():
        for line in key_file.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line.startswith("sk-") and "ant" not in line:
                return line
            if ": sk-" in line and "ant" not in line:
                return line.split(": ", 1)[1].strip()
    key = os.environ.get("OPENAI_API_KEY", "")
    if key:
        return key
    raise FileNotFoundError("No OpenAI key found.")

def load_anthropic_key(key_file: Path) -> str:
    if key_file.exists():
        for line in key_file.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line.startswith("sk-ant-"):
                return line
            if ": sk-ant-" in line:
                return line.split(": ", 1)[1].strip()
    key = os.environ.get("ANTHROPIC_API_KEY", "")
    if key:
        return key
    raise FileNotFoundError("No Anthropic key found.")

gpt_client    = OpenAI(api_key=load_openai_key(KEY_FILE))
claude_client = anthropic.Anthropic(api_key=load_anthropic_key(KEY_FILE))

# ---------------------------------------------------------------------------
# Build system prompt
# ---------------------------------------------------------------------------
def build_system_prompt(md_path: Path) -> str:
    text = md_path.read_text(encoding="utf-8")
    before = text.split("## Bill Information")[0].rstrip()
    output = "## Output" + text.split("## Output")[1].rstrip()
    return before + "\n\n" + output

SYSTEM_PROMPT = build_system_prompt(PROMPT_MD)

# ---------------------------------------------------------------------------
# Load step-1 flagged bills
# ---------------------------------------------------------------------------
step1_rows = {}
with open(STEP1_CSV, encoding="utf-8") as f:
    for r in csv.DictReader(f):
        if r["category_ids"].strip() in ("1", "1.0"):
            step1_rows[r["bill_id"]] = r

print(f"Step-1 flagged bills: {len(step1_rows)}")

# ---------------------------------------------------------------------------
# Load metadata
# ---------------------------------------------------------------------------
meta = {}
with open(META_CSV, encoding="utf-8") as f:
    for r in csv.DictReader(f):
        if r["bill_id"] in step1_rows:
            meta[r["bill_id"]] = r

print(f"Metadata found: {len(meta)} / {len(step1_rows)}")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def build_user_message(bill_id: str, m: dict) -> str:
    text_file = TEXT_DIR / f"{bill_id}_summary.txt"
    summary   = text_file.read_text(encoding="utf-8").strip() if text_file.exists() else m.get("summary", "")
    return (
        f"## Bill Information\n\n"
        f"**Official title:** {m.get('official_title', '')}\n"
        f"**Short title:** {m.get('short_title', '')}\n"
        f"**Policy area:** {m.get('policy_area', '')}\n"
        f"**Subjects:** {m.get('subjects', '')}\n"
        f"**Summary:**\n{summary}"
    )

def strip_code_fence(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        text = text.split("\n", 1)[1] if "\n" in text else text
        if text.endswith("```"):
            text = text[:-3].rstrip()
    return text

def parse_response(content: str) -> tuple[str, str, str, str]:
    parsed        = json.loads(content)
    cats          = parsed.get("categories", [])
    bill_type     = parsed.get("bill_type", "")
    bill_category = parsed.get("bill_category", "")
    reason        = parsed.get("reason", "")
    cats_str      = ";".join(str(c) for c in cats) if cats else ""
    return cats_str, bill_type, bill_category, reason

out_fields = [
    "bill_id", "official_title", "short_title", "policy_area", "subjects",
    "category_step1",   "reason_step1",
    "category_4o",      "bill_type_4o",      "bill_category_4o",      "reason_4o",
    "category_sonnet",  "bill_type_sonnet",  "bill_category_sonnet",  "reason_sonnet",
    "category_opus",    "bill_type_opus",    "bill_category_opus",    "reason_opus",
]

# ---------------------------------------------------------------------------
# Token / cost tracking (TRUE counts from API)
# ---------------------------------------------------------------------------
usage = {
    "gpt_input":    0,
    "gpt_output":   0,
    "sonnet_input": 0,
    "sonnet_output":0,
    "opus_input":   0,
    "opus_output":  0,
}

# ---------------------------------------------------------------------------
# Load checkpoint
# ---------------------------------------------------------------------------
done: dict[str, dict] = {}
if CHECKPOINT.exists():
    with open(CHECKPOINT, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            done[r["bill_id"]] = r
    print(f"Resumed from checkpoint: {len(done)} bills already done.")

results = list(done.values())
new_since_checkpoint = 0

# ---------------------------------------------------------------------------
# Main loop
# ---------------------------------------------------------------------------
bill_list = list(step1_rows.items())

for idx, (bill_id, s1_row) in enumerate(bill_list, 1):
    if bill_id in done:
        continue

    m = meta.get(bill_id, {})
    if not m:
        print(f"[{idx}/{len(bill_list)}] {bill_id} -- no metadata, skipping")
        continue

    step1_reason = s1_row.get("reason", "")
    print(f"\n[{idx}/{len(bill_list)}] {bill_id}")
    user_msg = build_user_message(bill_id, m)

    # --- GPT-4o ---
    gpt_cats = gpt_bt = gpt_bc = gpt_reason = "API error"
    for attempt in range(1, 4):
        try:
            resp = gpt_client.chat.completions.create(
                model=GPT_MODEL,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user",   "content": user_msg},
                ],
                temperature=0.0,
                max_tokens=400,
                response_format={"type": "json_object"},
            )
            gpt_cats, gpt_bt, gpt_bc, gpt_reason = parse_response(resp.choices[0].message.content)
            usage["gpt_input"]  += resp.usage.prompt_tokens
            usage["gpt_output"] += resp.usage.completion_tokens
            print(f"  GPT-4o:  [{gpt_cats or '-'}] [{gpt_bt or '-'}] [{gpt_bc or '-'}] -- {gpt_reason[:80]}")
            break
        except Exception as e:
            print(f"  GPT-4o retry {attempt}: {e}")
            time.sleep(2 ** attempt)

    # --- Claude Sonnet ---
    cl_cats = cl_bt = cl_bc = cl_reason = "API error"
    for attempt in range(1, 4):
        try:
            resp = claude_client.messages.create(
                model=SONNET_MODEL,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": user_msg}],
                temperature=0.0,
                max_tokens=400,
            )
            content = strip_code_fence(resp.content[0].text)
            cl_cats, cl_bt, cl_bc, cl_reason = parse_response(content)
            usage["sonnet_input"]  += resp.usage.input_tokens
            usage["sonnet_output"] += resp.usage.output_tokens
            print(f"  Sonnet:  [{cl_cats or '-'}] [{cl_bt or '-'}] [{cl_bc or '-'}] -- {cl_reason[:80]}")
            break
        except Exception as e:
            print(f"  Sonnet retry {attempt}: {e}")
            time.sleep(2 ** attempt)

    # --- Claude Opus ---
    op_cats = op_bt = op_bc = op_reason = "API error"
    for attempt in range(1, 4):
        try:
            resp = claude_client.messages.create(
                model=OPUS_MODEL,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": user_msg}],
                temperature=0.0,
                max_tokens=400,
            )
            content = strip_code_fence(resp.content[0].text)
            op_cats, op_bt, op_bc, op_reason = parse_response(content)
            usage["opus_input"]  += resp.usage.input_tokens
            usage["opus_output"] += resp.usage.output_tokens
            print(f"  Opus:    [{op_cats or '-'}] [{op_bt or '-'}] [{op_bc or '-'}] -- {op_reason[:80]}")
            break
        except Exception as e:
            print(f"  Opus retry {attempt}: {e}")
            time.sleep(2 ** attempt)

    row_out = {
        "bill_id":              bill_id,
        "official_title":       m.get("official_title", ""),
        "short_title":          m.get("short_title", ""),
        "policy_area":          m.get("policy_area", ""),
        "subjects":             m.get("subjects", ""),
        "category_step1":       s1_row.get("category_ids", ""),
        "reason_step1":         step1_reason,
        "category_4o":          gpt_cats,
        "bill_type_4o":         gpt_bt,
        "bill_category_4o":     gpt_bc,
        "reason_4o":            gpt_reason,
        "category_sonnet":      cl_cats,
        "bill_type_sonnet":     cl_bt,
        "bill_category_sonnet": cl_bc,
        "reason_sonnet":        cl_reason,
        "category_opus":        op_cats,
        "bill_type_opus":       op_bt,
        "bill_category_opus":   op_bc,
        "reason_opus":          op_reason,
    }
    results.append(row_out)
    new_since_checkpoint += 1

    if new_since_checkpoint >= CHECKPOINT_N:
        with open(CHECKPOINT, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=out_fields)
            writer.writeheader()
            writer.writerows(results)
        print(f"  [Checkpoint saved: {len(results)} rows]")
        new_since_checkpoint = 0

# ---------------------------------------------------------------------------
# Save final compare CSV
# ---------------------------------------------------------------------------
OUTPUT_CSV.parent.mkdir(exist_ok=True)
with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=out_fields)
    writer.writeheader()
    writer.writerows(results)

print(f"\nDone. {len(results)} rows saved to {OUTPUT_CSV}")

if CHECKPOINT.exists():
    CHECKPOINT.unlink()

# ---------------------------------------------------------------------------
# Summary table
# ---------------------------------------------------------------------------
def is_flagged(val):
    return str(val).strip() not in ("", "0", "0.0", "[]", "API error")

print("\n=== SUMMARY ===")
print(f"{'Bill':<18} {'Step1':>6} {'4o':>4} {'Sonnet':>7} {'Opus':>6}")
print("-" * 46)
for r in results:
    s = "Y" if is_flagged(r["category_step1"]) else "N"
    o = "Y" if is_flagged(r["category_4o"])    else "N"
    c = "Y" if is_flagged(r["category_sonnet"]) else "N"
    p = "Y" if is_flagged(r["category_opus"])   else "N"
    flag = "  *" if not (o == c == p) else ""
    print(f"{r['bill_id']:<18} {s:>6} {o:>4} {c:>7} {p:>6}{flag}")

s1_total   = sum(1 for r in results if is_flagged(r["category_step1"]))
o_total    = sum(1 for r in results if is_flagged(r["category_4o"]))
c_total    = sum(1 for r in results if is_flagged(r["category_sonnet"]))
p_total    = sum(1 for r in results if is_flagged(r["category_opus"]))
all3_agree = sum(1 for r in results
                 if is_flagged(r["category_4o"]) == is_flagged(r["category_sonnet"]) == is_flagged(r["category_opus"]))

print(f"\nAll 3 agree:    {all3_agree}/{len(results)}")
print(f"Step-1 flagged: {s1_total}")
print(f"4o flagged:     {o_total}")
print(f"Sonnet flagged: {c_total}")
print(f"Opus flagged:   {p_total}")

# ---------------------------------------------------------------------------
# Generate manual review xlsx (any model disagreement among the 3)
# ---------------------------------------------------------------------------
disagree_rows = [
    r for r in results
    if not (is_flagged(r["category_4o"]) == is_flagged(r["category_sonnet"]) == is_flagged(r["category_opus"]))
]

print(f"\nAny-model disagreement: {len(disagree_rows)} bills -> {REVIEW_XLS}")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Manual Review"

header_font         = Font(bold=True, color="FFFFFF")
header_fill_4o      = PatternFill("solid", fgColor="1F4E79")
header_fill_sonnet  = PatternFill("solid", fgColor="375623")
header_fill_opus    = PatternFill("solid", fgColor="5C3292")
header_fill_meta    = PatternFill("solid", fgColor="404040")
header_fill_manual  = PatternFill("solid", fgColor="7B2C2C")
fill_4o_yes         = PatternFill("solid", fgColor="BDD7EE")
fill_sonnet_yes     = PatternFill("solid", fgColor="C6E0B4")
fill_opus_yes       = PatternFill("solid", fgColor="D9C3F0")

columns = [
    ("bill_id",          "bill_id",               14, "meta"),
    ("official_title",   "official_title",         60, "meta"),
    ("short_title",      "short_title",            30, "meta"),
    ("policy_area",      "policy_area",            20, "meta"),
    ("step1_reason",     "reason_step1",           50, "meta"),
    ("4o_cat",           "category_4o",             7, "4o"),
    ("4o_type",          "bill_type_4o",           14, "4o"),
    ("4o_bill_cat",      "bill_category_4o",       12, "4o"),
    ("4o_reason",        "reason_4o",              55, "4o"),
    ("sonnet_cat",       "category_sonnet",         9, "sonnet"),
    ("sonnet_type",      "bill_type_sonnet",       14, "sonnet"),
    ("sonnet_bill_cat",  "bill_category_sonnet",   12, "sonnet"),
    ("sonnet_reason",    "reason_sonnet",           55, "sonnet"),
    ("opus_cat",         "category_opus",           8, "opus"),
    ("opus_type",        "bill_type_opus",         14, "opus"),
    ("opus_bill_cat",    "bill_category_opus",     12, "opus"),
    ("opus_reason",      "reason_opus",             55, "opus"),
    ("manual_decision",  None,                      16, "manual"),
    ("notes",            None,                      40, "manual"),
]

fill_map = {
    "meta":   header_fill_meta,
    "4o":     header_fill_4o,
    "sonnet": header_fill_sonnet,
    "opus":   header_fill_opus,
    "manual": header_fill_manual,
}

for col_idx, (header, _, width, group) in enumerate(columns, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = fill_map[group]
    cell.alignment = Alignment(wrap_text=False, vertical="center")
    ws.column_dimensions[cell.column_letter].width = width
ws.row_dimensions[1].height = 20

for row_idx, r in enumerate(disagree_rows, 2):
    f4o  = is_flagged(r["category_4o"])
    fson = is_flagged(r["category_sonnet"])
    fop  = is_flagged(r["category_opus"])
    for col_idx, (_, field, _, group) in enumerate(columns, 1):
        value = r.get(field, "") if field else ""
        cell  = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if group == "4o"     and f4o:  cell.fill = fill_4o_yes
        elif group == "sonnet" and fson: cell.fill = fill_sonnet_yes
        elif group == "opus"   and fop:  cell.fill = fill_opus_yes

ws.freeze_panes = "A2"
wb.save(REVIEW_XLS)
print(f"Manual review workbook saved: {REVIEW_XLS}")

# ---------------------------------------------------------------------------
# Write TRUE cost file
# ---------------------------------------------------------------------------
import datetime
now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

gpt_in_cost  = usage["gpt_input"]    / 1_000_000 *  2.50
gpt_out_cost = usage["gpt_output"]   / 1_000_000 * 10.00
son_in_cost  = usage["sonnet_input"] / 1_000_000 *  3.00
son_out_cost = usage["sonnet_output"]/ 1_000_000 * 15.00
op_in_cost   = usage["opus_input"]   / 1_000_000 * 15.00
op_out_cost  = usage["opus_output"]  / 1_000_000 * 75.00
total_cost   = gpt_in_cost + gpt_out_cost + son_in_cost + son_out_cost + op_in_cost + op_out_cost

cost_lines = f"""API Cost - compare_206bills_07_31_step2b.py
==========================================
Run completed: {now} (local time)
Script:        GPT_identify_bills/compare_206bills_07_31_step2b.py
Prompt:        HPC/prompt_step2_gpt.md
Input pool:    llm_bill_VT27_merged_2010sample5000_07_31_step1.csv
Note:          Step-1 reason NOT passed to LLMs (variant b)
Bills processed: {len(results)}

--- TRUE Token Counts (from API response.usage) ---

GPT-4o ({GPT_MODEL}):
  Input tokens:   {usage['gpt_input']:,}
  Output tokens:  {usage['gpt_output']:,}

Claude Sonnet ({SONNET_MODEL}):
  Input tokens:   {usage['sonnet_input']:,}
  Output tokens:  {usage['sonnet_output']:,}

Claude Opus ({OPUS_MODEL}):
  Input tokens:   {usage['opus_input']:,}
  Output tokens:  {usage['opus_output']:,}

--- Pricing Used ---

GPT-4o:           $2.50 / 1M input,   $10.00 / 1M output
Claude Sonnet:    $3.00 / 1M input,   $15.00 / 1M output
Claude Opus:     $15.00 / 1M input,   $75.00 / 1M output

--- Cost Breakdown ---

GPT-4o:
  Input cost:   ${gpt_in_cost:.4f}
  Output cost:  ${gpt_out_cost:.4f}
  Subtotal:     ${gpt_in_cost + gpt_out_cost:.4f}

Claude Sonnet ({SONNET_MODEL}):
  Input cost:   ${son_in_cost:.4f}
  Output cost:  ${son_out_cost:.4f}
  Subtotal:     ${son_in_cost + son_out_cost:.4f}

Claude Opus ({OPUS_MODEL}):
  Input cost:   ${op_in_cost:.4f}
  Output cost:  ${op_out_cost:.4f}
  Subtotal:     ${op_in_cost + op_out_cost:.4f}

TOTAL:          ${total_cost:.4f}
"""

COST_TXT.parent.mkdir(exist_ok=True)
COST_TXT.write_text(cost_lines, encoding="utf-8")
print(f"\nCost file saved: {COST_TXT}")
print(cost_lines)
